#!/usr/bin/env python3
"""One-time migration script: import existing sales from coupon_sales.xlsx
into the Neon Postgres database.

RUN (with DATABASE_URL set to your Neon connection string):

    Windows PowerShell:
        $env:DATABASE_URL="postgresql://user:pass@ep-xxx.neon.tech/db?sslmode=require"
        python migrate_to_pg.py

    Linux / macOS:
        DATABASE_URL="postgresql://user:pass@ep-xxx.neon.tech/db?sslmode=require" python migrate_to_pg.py

This script:
  1. Reads all rows from coupon_sales.xlsx (via openpyxl)
  2. Creates the sales + draw_results tables in Postgres if they don't exist
  3. Inserts every sale row into the Postgres sales table
  4. If a 'Lucky Draw' sheet exists in the Excel file, imports those too
  5. Prints a confirmation count

Safe to re-run: it clears the Postgres tables before inserting, so you
won't get duplicate rows.  Your local Excel file is NOT modified or
deleted — it stays as your offline backup.
"""

import os
import sys
from datetime import datetime

import openpyxl

import rohit


def main():
    if not os.environ.get("DATABASE_URL"):
        print("ERROR: DATABASE_URL environment variable is not set.")
        print("Set it to your Neon connection string and re-run.")
        print()
        print("Windows PowerShell:")
        print('  $env:DATABASE_URL="postgresql://user:pass@ep-xxx.neon.tech/db?sslmode=require"')
        print("  python migrate_to_pg.py")
        print()
        print("Linux / macOS:")
        print('  DATABASE_URL="postgresql://..." python migrate_to_pg.py')
        sys.exit(1)

    if not rohit._USE_POSTGRES:
        print("ERROR: rohit._USE_POSTGRES is False even though DATABASE_URL is set.")
        print("This shouldn't happen — check that the env var is correct.")
        sys.exit(1)

    print("=" * 60)
    print("MIGRATE: Excel -> Postgres")
    print("=" * 60)

    # Make sure the Postgres tables exist.
    rohit.init_db()
    print("[1/4] Postgres tables created (if not already present).")

    # Read the Excel file.
    if not rohit.SALES_FILE.exists():
        print(f"ERROR: {rohit.SALES_FILE} not found. Nothing to migrate.")
        sys.exit(1)

    wb = openpyxl.load_workbook(rohit.SALES_FILE, read_only=True)
    ws = wb.active
    excel_rows = list(ws.iter_rows(min_row=2, values_only=True))
    excel_rows = [r for r in excel_rows if any(c is not None for c in r)]
    wb.close()

    print(f"[2/4] Read {len(excel_rows)} sale rows from {rohit.SALES_FILE.name}.")

    # Check for a Lucky Draw sheet.
    draw_rows = []
    wb2 = openpyxl.load_workbook(rohit.SALES_FILE, read_only=True)
    if rohit.DRAW_SHEET in wb2.sheetnames:
        dws = wb2[rohit.DRAW_SHEET]
        draw_rows = list(dws.iter_rows(min_row=2, values_only=True))
        draw_rows = [r for r in draw_rows if any(c is not None for c in r)]
    wb2.close()
    print(f"[3/4] Read {len(draw_rows)} draw result rows from Excel.")

    # Clear the Postgres tables (safe re-run).
    conn = rohit._pg_conn()
    try:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM sales")
            cur.execute("DELETE FROM draw_results")

            # Insert sale rows.
            for r in excel_rows:
                # Excel columns: S.No, Name, Phone, Address, Start, End, Qty,
                # Date, Type, Set Size, Amount
                sno = r[0] if len(r) > 0 else None
                name = r[1] if len(r) > 1 else None
                phone = r[2] if len(r) > 2 else None
                address = r[3] if len(r) > 3 else None
                start = r[4] if len(r) > 4 else None
                end = r[5] if len(r) > 5 else None
                qty = r[6] if len(r) > 6 else None
                date = r[7] if len(r) > 7 else None
                stype = r[8] if len(r) > 8 and r[8] else "SALE"
                set_sz = r[9] if len(r) > 9 else None
                amount = r[10] if len(r) > 10 else None

                if start is None or end is None:
                    continue

                # Normalize set_size: blank -> NULL for Postgres.
                if set_sz == "" or set_sz is None:
                    set_sz = None
                else:
                    try:
                        set_sz = int(set_sz)
                    except (ValueError, TypeError):
                        set_sz = None

                cur.execute("""
                    INSERT INTO sales
                        (sno, name, phone, address, start_no, end_no, qty,
                         date_sold, sale_type, set_size, amount)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (sno, name, phone, address, int(start), int(end),
                      qty, str(date) if date else None, stype, set_sz,
                      amount))
            sales_count = len(excel_rows)

            # Insert draw result rows (if any).
            for r in draw_rows:
                # Draw columns: Prize, Gift, CouponNo, SetRange, Buyer,
                # Phone, Type, DrawnAt
                prize = r[0] if len(r) > 0 else None
                gift = r[1] if len(r) > 1 else None
                coupon_no = r[2] if len(r) > 2 else None
                set_range = r[3] if len(r) > 3 else None
                buyer = r[4] if len(r) > 4 else None
                phone = r[5] if len(r) > 5 else None
                stype = r[6] if len(r) > 6 else None
                drawn_at = r[7] if len(r) > 7 else None

                # Parse set_range "0301-0310" -> (start, end)
                range_start = None
                range_end = None
                if set_range and "-" in str(set_range):
                    parts = str(set_range).split("-")
                    try:
                        range_start = int(parts[0])
                        range_end = int(parts[1])
                    except (ValueError, IndexError):
                        pass

                cur.execute("""
                    INSERT INTO draw_results
                        (prize, gift, coupon_no, range_start, range_end,
                         buyer, phone, sale_type, drawn_at)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, (prize, gift, coupon_no, range_start, range_end,
                      buyer, phone, stype, str(drawn_at) if drawn_at else None))
            draw_count = len(draw_rows)

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        rohit._pg_release(conn)

    print(f"[4/4] Inserted {sales_count} sale(s) + {draw_count} draw result(s) into Postgres.")
    print()
    print("=" * 60)
    print("MIGRATION COMPLETE")
    print("=" * 60)
    print(f"Sales imported : {sales_count}")
    print(f"Draw results   : {draw_count}")
    print(f"Your Excel file is untouched and remains as a local backup.")
    print()
    print("Next step: deploy to Render (or it will auto-build on push),")
    print("then open your Render URL — all sales will be visible.")


if __name__ == "__main__":
    main()