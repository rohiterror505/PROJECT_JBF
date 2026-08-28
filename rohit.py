#!/usr/bin/env python3
"""
JAI BHADRA FOUNDATION
LUCKY DRAW COUPON GENERATOR

Design:
- Professional cream / maroon / gold ticket
- Left and right deity photos
- Coupon number 0001 -> 0005 for testing
- QR code containing the coupon number
- Draw date: 06 September 2026
- Day: Sunday
- 10 main prizes
- 4 consolation prizes
- No prices printed on the coupon
- Footer: https://jaibhadra.org/

INTERACTIVE SALE MODE:
- The script asks for a buyer's name, phone, address, and a start/end
  coupon number range.
- For a single coupon, enter the same number as both start and end
  (e.g. 5 and 5 -> coupon 0005).
- For multiple coupons, enter a range (e.g. 1 and 10 -> 0001..0010).
- Each sale is appended to coupon_sales.xlsx (S.No, Name, Phone,
  Address, Start, End, Qty, Date Sold) so you can track which coupon
  went to whom.
- A coupon range that overlaps an already-sold range is rejected to
  prevent double-selling.
- Type 'q' at any prompt to finish.

FILES REQUIRED IN THE SAME FOLDER:
    rohit.py
    left_photo.png
    right_photo.png

INSTALL:
    python -m pip install "qrcode[pil]" pillow openpyxl

RUN:
    python rohit.py

OUTPUT:
    lucky_draw_coupons/
        coupon_0001.png
        coupon_0002.png
        ...
    coupon_sales.xlsx        (sales tracker, auto-created)
"""

from datetime import datetime
from pathlib import Path
import shutil
import zipfile

import openpyxl
import qrcode
from PIL import Image, ImageDraw, ImageFont, ImageOps


# ============================================================
# SETTINGS
# ============================================================

MAX_COUPON = 5000

# ------------------------------------------------------------
# DONATION PRICING
# A "New Sale" (named buyer) is always charged flat ₹100 per coupon.
# A "Physical Set" strip prints the TOTAL donation for the whole set,
# so a Set of 10 prints ₹1000, a Set of 5 prints ₹500, a Set of 1 prints ₹100.
# The per-coupon rate is constant (₹100); only the displayed/recorded
# amount differs because a set strip represents one donation for the
# whole strip.
# ------------------------------------------------------------
PRICE_PER_COUPON = 100

# Map a physical set size -> total donation amount for that set strip.
SET_PRICES = {1: 100, 5: 500, 10: 1000}


def price_for_set_size(set_size):
    """Return the total donation amount for a physical set of the given
    size.  Falls back to set_size * PRICE_PER_COUPON for any size not
    explicitly listed in SET_PRICES."""
    return SET_PRICES.get(set_size, set_size * PRICE_PER_COUPON)


BASE_DIR = Path(__file__).resolve().parent

LEFT_PHOTO = BASE_DIR / "left_photo.png"
RIGHT_PHOTO = BASE_DIR / "right_photo.png"
LOGO = BASE_DIR / "images.jpg"

OUTPUT_DIR = BASE_DIR / "lucky_draw_coupons"
SALES_FILE = BASE_DIR / "coupon_sales.xlsx"

# In-memory cache of the sales workbook rows so the UI does not have to
# re-open and re-parse coupon_sales.xlsx on every refresh / tab switch
# (which is what caused the "Not Responding" freezes, especially when the
# file lives on OneDrive).  The cache is invalidated by every function that
# mutates the workbook (record_sale, delete_sale, delete_all_*,
# delete_physical_*).  Callers that only READ (list_sales, get_sold_ranges,
# is_already_sold) use the cache and fall back to a live load on miss.
_sales_cache = {"rows": None, "mtime": None}

# Whether ensure_sales_file has already validated the existing workbook
# in this process.  Once True, subsequent ensure_sales_file() calls skip
# re-opening the file (the happy path) to avoid the per-refresh lag.
_ensure_done = False


def _sales_mtime():
    try:
        return SALES_FILE.stat().st_mtime
    except OSError:
        return None


def _invalidate_sales_cache():
    _sales_cache["rows"] = None
    _sales_cache["mtime"] = None
    # Force ensure_sales_file to re-validate the workbook on next call,
    # since a write may have just changed the file on disk.
    global _ensure_done
    _ensure_done = False


def _get_sales_rows():
    """Return the list of sales rows (excluding header), using an in-memory
    cache keyed on the file's mtime.  Reloads from disk only when the file
    has changed.  Raises RuntimeError if the file is locked."""
    mtime = _sales_mtime()
    if _sales_cache["rows"] is not None and _sales_cache["mtime"] == mtime:
        return _sales_cache["rows"]

    rows = _load_sales_rows_from_disk()
    _sales_cache["rows"] = rows
    _sales_cache["mtime"] = mtime
    return rows


def _load_sales_rows_from_disk():
    """Open coupon_sales.xlsx and return the non-empty data rows."""
    wb = _load_sales_workbook(read_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        rows = [r for r in rows if any(c is not None for c in r)]
    finally:
        wb.close()
        # Clean up the temp copy if _load_sales_workbook made one.
        tmp = getattr(wb, "_jbf_tmp_path", None)
        if tmp:
            try:
                Path(tmp).unlink(missing_ok=True)
            except Exception:
                pass
    return rows

# High-resolution portrait ticket (phone view).
WIDTH = 1200
HEIGHT = 2140

DRAW_DATE = "06 SEPTEMBER 2026"
DRAW_DAY = "SUNDAY"
DRAW_LINE = f"{DRAW_DAY}, {DRAW_DATE}"
VENUE = "Golden Star Banquet (Petals Hall), Swarn Jayanti Park (Japanese Park)"
WEBSITE = "www.jaibhadra.org"


# ============================================================
# PRIZES
# No prices are printed.
# ============================================================

MAIN_PRIZES = [
    ("1st Prize", "\U0001F4BB HP Professional 15 Laptop (i3 14th Gen, 8GB, 512GB SSD)"),
    ("2nd Prize", "\U0001F9FE Whirlpool Refrigerator 184L"),
    ("3rd Prize", "\U0001F300 Whirlpool Washing Machine (fully automatic)"),
    ("4th Prize", "\U0001F4F1 Realme Narzo 80 Lite Smartphone"),
    ("5th Prize", "\U0001F4FA Realme TechLife 32 inch Smart TV"),
    ("6th Prize", "\U0001F373 Panasonic 20L Microwave Oven"),
    ("7th Prize", "\U0001F525 Philips NA120/00 Air Fryer"),
    ("8th Prize", "\U0001F96A Philips HD2288/00 Sandwich Maker"),
    ("9th Prize", "\U0001F9F5 Havells Mixer Grinder"),
    ("10th Prize", "\U0001F6C5 Aristocrat Trolley Bag"),
]

CONSOLATION_PRIZES = [
    ("A", "\U0001F37D Borosil Dinner Set", "(10)"),
    ("B", "\U0001F375 Milton 1.8L Electric Kettle", "(10)"),
    ("C", "\U0001F3A7 Boat Earphones", "(10)"),
    ("D", "\U0001F4A8 Havells Electric Press / Iron", "(10)"),
]


# ============================================================
# COLORS
# ============================================================

CREAM = "#FFF9E8"
CREAM_2 = "#FFFDF6"
MAROON = "#E85D04"
MAROON_DARK = "#C44404"
NAVY = "#E85D04"
NAVY_DARK = "#C44404"

GOLD = "#C99B2E"
GOLD_LIGHT = "#F3D264"
GOLD_PALE = "#FFF1BD"
GOLD_LINE = "#D5AF50"
GOLD_WHITE = "#FBF3D9"

BLACK = "#151515"
GREY = "#666666"
WHITE = "#FFFFFF"
TABLE_LINE = "#D8BD79"


# ============================================================
# FONTS
# ============================================================

def get_font(size, bold=False, italic=False):
    if italic:
        candidates = [
            r"C:\Windows\Fonts\georgiai.ttf",
            r"C:\Windows\Fonts\timesi.ttf",
            r"C:\Windows\Fonts\cambriai.ttf",
            r"C:\Windows\Fonts\ariali.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSerif-Italic.ttf",
        ]
    elif bold:
        candidates = [
            r"C:\Windows\Fonts\georgiab.ttf",
            r"C:\Windows\Fonts\cambriaz.ttf",
            r"C:\Windows\Fonts\timesbd.ttf",
            r"C:\Windows\Fonts\arialbd.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSerif-Bold.ttf",
        ]
    else:
        candidates = [
            r"C:\Windows\Fonts\georgia.ttf",
            r"C:\Windows\Fonts\cambria.ttf",
            r"C:\Windows\Fonts\times.ttf",
            r"C:\Windows\Fonts\arial.ttf",
            "/usr/share/fonts/truetype/dejavu/DejaVuSerif.ttf",
        ]

    for font_path in candidates:
        if Path(font_path).exists():
            return ImageFont.truetype(font_path, size)

    return ImageFont.load_default()


def get_emoji_font(size):
    candidates = [
        r"C:\Windows\Fonts\seguiemj.ttf",
        r"C:\Windows\Fonts\seguisym.ttf",
        "/usr/share/fonts/truetype/noto/NotoColorEmoji.ttf",
    ]

    for font_path in candidates:
        if Path(font_path).exists():
            return ImageFont.truetype(font_path, size)

    return get_font(size)


def fit_font(draw, text, max_width, start_size, bold=False):
    size = start_size

    while size >= 10:
        font = get_font(size, bold=bold)

        box = draw.textbbox((0, 0), text, font=font)
        width = box[2] - box[0]

        if width <= max_width:
            return font

        size -= 2

    return get_font(10, bold=bold)


# ============================================================
# TEXT HELPERS
# ============================================================

def text_size(draw, text, font):
    box = draw.textbbox((0, 0), text, font=font)
    return box[2] - box[0], box[3] - box[1]


def draw_center(draw, text, box, font, fill):
    x1, y1, x2, y2 = box

    tw, th = text_size(draw, text, font)

    x = x1 + ((x2 - x1) - tw) / 2
    y = y1 + ((y2 - y1) - th) / 2

    draw.text(
        (int(x), int(y)),
        text,
        font=font,
        fill=fill,
    )


def draw_center_y(draw, text, y, left, right, font, fill):
    tw, _ = text_size(draw, text, font)

    x = left + ((right - left) - tw) / 2

    draw.text(
        (int(x), int(y)),
        text,
        font=font,
        fill=fill,
    )


def draw_mixed_center(draw, text, box, font, emoji_font, fill):
    """
    Draw text that may contain emoji. Emoji glyphs are rendered with the
    emoji font, the rest with the regular font, all centered in `box`.
    """
    x1, y1, x2, y2 = box

    # Split into runs: (is_emoji, chunk)
    runs = []
    buf = ""
    in_emoji = False

    for ch in text:
        is_emoji = ord(ch) > 0x1F000 or ch in "\u2600\u2601\u260E\u2615\u2665\u2764\u2B50\u2728"

        if not runs and not buf:
            in_emoji = is_emoji

        if is_emoji == in_emoji:
            buf += ch
        else:
            runs.append((in_emoji, buf))
            buf = ch
            in_emoji = is_emoji

    if buf:
        runs.append((in_emoji, buf))

    # Measure total width.
    total_w = 0
    for is_emoji, chunk in runs:
        f = emoji_font if is_emoji else font
        tw, _ = text_size(draw, chunk, f)
        total_w += tw

    x = x1 + ((x2 - x1) - total_w) / 2
    y = y1 + ((y2 - y1) - (font.size)) / 2

    for is_emoji, chunk in runs:
        f = emoji_font if is_emoji else font
        tw, _ = text_size(draw, chunk, f)
        draw.text((int(x), int(y)), chunk, font=f, fill=fill)
        x += tw


def draw_mixed_left(draw, text, box, font, emoji_font, fill, padding=15):
    """
    Draw text that may contain emoji, left-aligned in `box` with padding.
    """
    x1, y1, x2, y2 = box

    # Split into runs: (is_emoji, chunk)
    runs = []
    buf = ""
    in_emoji = False

    for ch in text:
        is_emoji = ord(ch) > 0x1F000 or ch in "\u2600\u2601\u260E\u2615\u2665\u2764\u2B50\u2728"

        if not runs and not buf:
            in_emoji = is_emoji

        if is_emoji == in_emoji:
            buf += ch
        else:
            runs.append((in_emoji, buf))
            buf = ch
            in_emoji = is_emoji

    if buf:
        runs.append((in_emoji, buf))

    x = x1 + padding
    y = y1 + ((y2 - y1) - (font.size)) / 2

    for is_emoji, chunk in runs:
        f = emoji_font if is_emoji else font
        tw, _ = text_size(draw, chunk, f)
        draw.text((int(x), int(y)), chunk, font=f, fill=fill)
        x += tw


# ============================================================
# DECORATIVE FLOURISH
# ============================================================

def flourish(draw, cx, cy, scale=1.0, color=GOLD):
    """
    Small symmetrical gold ornament.
    Uses lines/ellipses only, so it is safe on Pillow versions.
    """

    w = int(70 * scale)
    h = int(20 * scale)

    draw.arc(
        (
            cx - w,
            cy - h,
            cx,
            cy + h,
        ),
        200,
        350,
        fill=color,
        width=max(1, int(3 * scale)),
    )

    draw.arc(
        (
            cx,
            cy - h,
            cx + w,
            cy + h,
        ),
        190,
        340,
        fill=color,
        width=max(1, int(3 * scale)),
    )

    draw.ellipse(
        (
            cx - 5 * scale,
            cy - 5 * scale,
            cx + 5 * scale,
            cy + 5 * scale,
        ),
        fill=color,
    )


def horizontal_ornament(draw, x1, x2, y, color=GOLD):
    mid = (x1 + x2) // 2

    draw.line(
        (x1, y, mid - 35, y),
        fill=color,
        width=3,
    )

    draw.line(
        (mid + 35, y, x2, y),
        fill=color,
        width=3,
    )

    flourish(
        draw,
        mid,
        y,
        scale=0.65,
        color=color,
    )


def draw_lotus(draw, cx, cy, size, color):
    """
    Draw a simple lotus flower with a center and 6 petals.
    Festive Indian floral accent.
    """
    import math

    # Petals as small ellipses radiating from the center.
    petal_w = int(size * 0.55)
    petal_h = int(size * 0.85)
    for angle in range(0, 360, 60):
        rad = math.radians(angle)
        dx = int((size * 0.35) * math.cos(rad))
        dy = int((size * 0.35) * math.sin(rad))
        draw.ellipse(
            (
                cx + dx - petal_w // 2,
                cy + dy - petal_h // 2,
                cx + dx + petal_w // 2,
                cy + dy + petal_h // 2,
            ),
            outline=color,
            width=max(1, int(size * 0.08)),
        )

    # Center circle.
    r = int(size * 0.18)
    draw.ellipse((cx - r, cy - r, cx + r, cy + r), fill=color)


def draw_marigold(draw, cx, cy, size, color):
    """
    Draw a circular marigold-like flower with many small petals.
    """
    import math

    petals = 12
    outer_r = int(size * 0.45)
    inner_r = int(size * 0.22)

    for i in range(petals):
        angle = math.radians(i * (360 / petals))
        px = cx + int(outer_r * math.cos(angle))
        py = cy + int(outer_r * math.sin(angle))
        r = int(size * 0.18)
        draw.ellipse((px - r, py - r, px + r, py + r), outline=color, width=1)

    draw.ellipse(
        (cx - inner_r, cy - inner_r, cx + inner_r, cy + inner_r),
        fill=color,
    )


def draw_leaf(draw, cx, cy, size, angle, color):
    """
    Draw a simple paisley/teardrop leaf at the given angle.
    """
    import math

    rad = math.radians(angle)
    # Two endpoints of the leaf.
    tip_x = cx + int(size * 0.6 * math.cos(rad))
    tip_y = cy + int(size * 0.6 * math.sin(rad))
    base_x = cx - int(size * 0.3 * math.cos(rad))
    base_y = cy - int(size * 0.3 * math.sin(rad))

    # Control points for a wider teardrop.
    perp_x = -math.sin(rad)
    perp_y = math.cos(rad)
    ctrl1_x = cx + int(size * 0.35 * perp_x)
    ctrl1_y = cy + int(size * 0.35 * perp_y)
    ctrl2_x = cx - int(size * 0.35 * perp_x)
    ctrl2_y = cy - int(size * 0.35 * perp_y)

    # Approximate with a filled polygon.
    pts = [
        (tip_x, tip_y),
        (ctrl1_x, ctrl1_y),
        (base_x, base_y),
        (ctrl2_x, ctrl2_y),
    ]
    draw.polygon(pts, fill=color)


def draw_floral_corner(draw, cx, cy, scale, color):
    """
    A medium-complexity corner floral motif (lotus + two leaves).
    """
    draw_lotus(draw, cx, cy, int(24 * scale), color)
    draw_leaf(draw, cx, cy, int(22 * scale), 45, color)
    draw_leaf(draw, cx, cy, int(22 * scale), -45, color)


# ============================================================
# BORDER
# ============================================================

def draw_ticket_border(draw):
    # Outer ticket border (maroon).
    draw.rounded_rectangle(
        (12, 12, WIDTH - 12, HEIGHT - 12),
        radius=38,
        fill=CREAM,
        outline=MAROON_DARK,
        width=6,
    )

    # Single gold inner border.
    draw.rounded_rectangle(
        (28, 28, WIDTH - 28, HEIGHT - 28),
        radius=30,
        outline=GOLD,
        width=2,
    )

    # Decorative gold corner accents (clean L-shapes).
    length = 46
    offset = 52

    for x, y, sx, sy in [
        (offset, offset, 1, 1),
        (WIDTH - offset, offset, -1, 1),
        (offset, HEIGHT - offset, 1, -1),
        (WIDTH - offset, HEIGHT - offset, -1, -1),
    ]:
        draw.line(
            (x, y, x + sx * length, y),
            fill=GOLD,
            width=3,
        )

        draw.line(
            (x, y, x, y + sy * length),
            fill=GOLD,
            width=3,
        )

        # Small diamond at the corner tip.
        d = 5
        draw.polygon(
            [
                (x, y - d),
                (x + d, y),
                (x, y + d),
                (x - d, y),
            ],
            fill=GOLD,
        )

    # Gold diamonds at the midpoints of the top and bottom inner border edges.
    for y_edge in (28, HEIGHT - 28):
        for x in [WIDTH // 2]:
            d = 5
            draw.polygon(
                [
                    (x, y_edge - d),
                    (x + d, y_edge),
                    (x, y_edge + d),
                    (x - d, y_edge),
                ],
                fill=GOLD,
            )

    # Gold diamonds at the midpoints of the left and right inner border edges.
    for x_edge in (28, WIDTH - 28):
        for y in [HEIGHT // 2]:
            d = 5
            draw.polygon(
                [
                    (x_edge - d, y),
                    (x_edge, y + d),
                    (x_edge + d, y),
                    (x_edge, y - d),
                ],
                fill=GOLD,
            )


def draw_text_watermark(draw):
    """No text watermark is drawn on the ticket."""
    return


def draw_stars(draw, cx, cy, count=3, radius=10, color=GOLD):
    """Small 5-point gold stars around a center point."""
    import math

    for i in range(count):
        angle = i * (360 / count) - 90
        sx = cx + radius * 2.2 * math.cos(math.radians(angle))
        sy = cy + radius * 2.2 * math.sin(math.radians(angle))

        # 5-point star polygon.
        pts = []
        for k in range(10):
            r = radius if k % 2 == 0 else radius * 0.45
            a = math.radians(-90 + k * 36)
            pts.append((sx + r * math.cos(a), sy + r * math.sin(a)))

        draw.polygon(pts, fill=color)


# ============================================================
# HEADER
# ============================================================

def draw_header(draw):
    # Header between the two photos.
    header = (280, 60, WIDTH - 280, 260)
    draw.rounded_rectangle(header, radius=20, fill=MAROON, outline=GOLD, width=3)
    draw.rounded_rectangle((290, 70, WIDTH - 290, 250), radius=14, outline=GOLD_LINE, width=1)

    for cx, cy in [(280, 60), (WIDTH - 280, 60), (280, 260), (WIDTH - 280, 260)]:
        d = 5
        draw.polygon([(cx, cy - d), (cx + d, cy), (cx, cy + d), (cx - d, cy)], fill=GOLD)

    heading = "JAIBHADRA FOUNDATION"
    draw_center(draw, heading, (290, 80, WIDTH - 290, 160),
                fit_font(draw, heading, WIDTH - 580, 40, bold=True), GOLD_LIGHT)

    draw.line((340, 175, WIDTH - 340, 175), fill=GOLD, width=2)
    for dx in (-1, 1):
        cx = (280 + WIDTH - 280) // 2 + dx * ((WIDTH - 280 - 280) // 2 - 60)
        d = 4
        draw.polygon([(cx, 175 - d), (cx + d, 175), (cx, 175 + d), (cx - d, 175)], fill=GOLD)

    draw_center(draw, "LUCKY DRAW COUPON", (290, 185, WIDTH - 290, 245),
                get_font(30, bold=True), CREAM)


# ============================================================
# PHOTO
# ============================================================

def load_photo(path, max_size):
    if not path.exists():
        print(f"WARNING: {path.name} not found.")
        return None

    try:
        img = Image.open(path).convert("RGBA")

        # CONTAIN = never crop the source image.
        return ImageOps.contain(
            img,
            max_size,
            method=Image.Resampling.LANCZOS,
        )

    except Exception as exc:
        print(f"WARNING: Could not read {path.name}: {exc}")
        return None


def load_photo_fixed_height(path, frame_w, target_h):
    """Scale photo to an exact target height (width scales proportionally).
    Never crops. Returns the scaled RGBA image."""
    if not path.exists():
        print(f"WARNING: {path.name} not found.")
        return None

    try:
        img = Image.open(path).convert("RGBA")
        orig_w, orig_h = img.size

        scale = target_h / orig_h
        new_w = int(orig_w * scale)

        return img.resize(
            (new_w, int(target_h)),
            Image.Resampling.LANCZOS,
        )

    except Exception as exc:
        print(f"WARNING: Could not read {path.name}: {exc}")
        return None


def compute_common_photo_height(frame_w, frame_h, paths):
    """Find a height that fits every photo (scaled to that height) within frame_w.
    Returns that height so all photos end up the same size without cropping."""
    heights = []
    for p in paths:
        if not p.exists():
            continue
        try:
            w, h = Image.open(p).size
            # Height when this photo's width equals frame_w.
            heights.append(frame_w * h / w)
        except Exception:
            pass

    if not heights:
        return frame_h

    # Use the smallest so the widest-proportioned photo still fits.
    return int(min(heights))


def paste_photo(canvas, photo, frame):
    if photo is None:
        return

    x1, y1, x2, y2 = frame

    fw = x2 - x1
    fh = y2 - y1

    px = x1 + (fw - photo.width) // 2
    py = y1 + (fh - photo.height) // 2

    mask = Image.new(
        "L",
        photo.size,
        0,
    )

    mask_draw = ImageDraw.Draw(mask)

    mask_draw.rounded_rectangle(
        (
            0,
            0,
            photo.width - 1,
            photo.height - 1,
        ),
        radius=24,
        fill=255,
    )

    canvas.paste(
        photo,
        (px, py),
        mask,
    )


def draw_photo_card(draw, canvas, photo_path, frame):
    x1, y1, x2, y2 = frame

    # Cream backing (no outline).
    draw.rounded_rectangle(
        frame,
        radius=20,
        fill=CREAM,
    )

    photo = load_photo(
        photo_path,
        (x2 - x1, y2 - y1),
    )

    if photo is None:
        return

    # Center horizontally and vertically within the frame.
    px = x1 + ((x2 - x1) - photo.width) // 2
    py = y1 + ((y2 - y1) - photo.height) // 2

    mask = Image.new(
        "L",
        photo.size,
        0,
    )

    mask_draw = ImageDraw.Draw(mask)

    mask_draw.rounded_rectangle(
        (
            0,
            0,
            photo.width - 1,
            photo.height - 1,
        ),
        radius=20,
        fill=255,
    )

    canvas.paste(
        photo,
        (px, py),
        mask,
    )


# ============================================================
# COUPON ID + QR
# ============================================================

def draw_coupon_center(draw, canvas, start, end, buyer=None, phone=None, address=None, amount=None, sample=False):
    is_range = start != end
    qty = end - start + 1

    # Donation amount printed on this strip.  For a physical set the amount
    # is the TOTAL for the whole strip (e.g. Set of 10 -> ₹1000).  For a
    # single New-Sale coupon it is the per-coupon rate (₹100).  When amount
    # is None we fall back to the per-coupon rate * qty so existing callers
    # that omit it still get the right number.
    if amount is None:
        amount = qty * PRICE_PER_COUPON

    card = (60, 300, WIDTH - 60, 700)
    draw.rounded_rectangle(card, radius=18, fill=WHITE, outline=GOLD, width=3)
    draw.rounded_rectangle((70, 310, WIDTH - 70, 690), radius=12, outline=GOLD_LINE, width=1)

    for cx, cy in [(60, 300), (WIDTH - 60, 300), (60, 700), (WIDTH - 60, 700)]:
        d = 5
        draw.polygon([(cx, cy - d), (cx + d, cy), (cx, cy + d), (cx - d, cy)], fill=GOLD)

    # QR code on the left side — contains buyer details so they show when scanned.
    if is_range:
        coupon_line = f"SAMPLE COUPONS: {start:04d}-{end:04d} ({qty} coupons)" if sample else f"COUPONS: {start:04d}-{end:04d} ({qty} coupons)"
    else:
        coupon_line = f"SAMPLE COUPON NUMBER: {start:04d}" if sample else f"COUPON NUMBER: {start:04d}"

    qr_lines = [
        "JAIBHADRA FOUNDATION",
        "LUCKY DRAW COUPON",
        coupon_line,
    ]
    if buyer:
        qr_lines.append(f"NAME: {buyer}")
    if phone:
        qr_lines.append(f"PHONE: {phone}")
    if address:
        qr_lines.append(f"ADDRESS: {address}")

    qr_data = "\n".join(qr_lines)
    qr = qrcode.QRCode(
        version=None,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=8, border=4,
    )
    qr.add_data(qr_data)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color="black", back_color="white").convert("RGB")

    QR_SIZE = 200
    qr_img = qr_img.resize((QR_SIZE, QR_SIZE), Image.Resampling.LANCZOS)

    qr_x = 110
    qr_y = 350
    border_pad = 12

    draw.rounded_rectangle(
        (qr_x - border_pad, qr_y - border_pad,
         qr_x + QR_SIZE + border_pad, qr_y + QR_SIZE + border_pad),
        radius=10, fill=WHITE, outline=GOLD, width=3,
    )
    canvas.paste(qr_img, (qr_x, qr_y))

    # Scan text below QR.
    draw_center(draw, "SCAN TO",
                (qr_x - 20, qr_y + QR_SIZE + border_pad + 15, qr_x + QR_SIZE + 20, qr_y + QR_SIZE + border_pad + 45),
                get_font(18, bold=True), MAROON)
    draw_center(draw, "VERIFY YOUR COUPON",
                (qr_x - 20, qr_y + QR_SIZE + border_pad + 45, qr_x + QR_SIZE + 20, qr_y + QR_SIZE + border_pad + 75),
                get_font(18, bold=True), MAROON)

    # Coupon number hero panel on the right side.
    panel = (450, 340, WIDTH - 110, 490)
    draw.rounded_rectangle(panel, radius=24, fill=MAROON, outline=GOLD, width=4)
    draw.rounded_rectangle((460, 350, WIDTH - 120, 480), radius=18, outline=GOLD_LINE, width=1)

    label_text = ("SAMPLE COUPON NUMBERS" if is_range else "SAMPLE COUPON NUMBER") if sample else ("COUPON NUMBERS" if is_range else "COUPON NUMBER")
    draw_center(draw, label_text, (460, 356, WIDTH - 120, 395),
                get_font(28, bold=True), GOLD_LIGHT)

    num_text = f"{start:04d}-{end:04d}({qty})" if is_range else f"{start:04d}"
    num_font = fit_font(draw, num_text, WIDTH - 120 - 460 - 60, 56, bold=True)
    draw_center(draw, num_text, (460, 395, WIDTH - 120, 475), num_font, GOLD_LIGHT)

    # Donation strip on the right side below number panel (10px gap).
    donation_strip = (450, 500, WIDTH - 110, 560)
    draw.rounded_rectangle(donation_strip, radius=16, fill=CREAM_2, outline=GOLD, width=2)

    donation_text = f"DONATION AMOUNT: \u20B9{amount}/-"
    draw_center(draw, donation_text,
                (460, 506, WIDTH - 120, 554),
                fit_font(draw, donation_text, WIDTH - 120 - 460 - 40, 28, bold=True),
                MAROON_DARK)

    # Total prizes strip below donation amount (10px gap).
    prizes_strip = (450, 570, WIDTH - 110, 640)
    draw.rounded_rectangle(prizes_strip, radius=16, fill=MAROON, outline=GOLD, width=2)

    draw_center(draw, "TOTAL: 50 PRIZES TO BE WON",
                (460, 576, WIDTH - 120, 634),
                fit_font(draw, "TOTAL: 50 PRIZES TO BE WON", WIDTH - 120 - 460 - 40, 26, bold=True),
                GOLD_LIGHT)


# ============================================================
# DATE / DAY
# ============================================================

def draw_date_box(draw):
    date_bar = (100, 715, WIDTH - 100, 915)
    draw.rounded_rectangle(date_bar, radius=14, fill=WHITE, outline=GOLD, width=3)
    draw.rounded_rectangle((110, 725, WIDTH - 110, 905), radius=10, outline=GOLD_LINE, width=1)

    festival_line = "Bhagwan Shree Balbhadra Jayanti & Bhagwan Shree Sahastrajun Pujnotsaw"
    fest_font = fit_font(draw, festival_line, WIDTH - 240 - 40, 28, bold=True)
    draw_center(draw, festival_line, (120, 735, WIDTH - 120, 775), fest_font, MAROON)

    draw_center(draw, DRAW_LINE, (120, 785, WIDTH - 120, 817),
                get_font(30, bold=True), MAROON_DARK)

    draw.line((280, 825, WIDTH - 280, 825), fill=GOLD, width=2)
    for cx in (280, WIDTH - 280):
        d = 4
        draw.polygon([(cx, 825 - d), (cx + d, 825), (cx, 825 + d), (cx - d, 825)], fill=GOLD)

    venue_font = fit_font(draw, VENUE, WIDTH - 240 - 40, 24, bold=True)
    draw_center(draw, VENUE, (120, 835, WIDTH - 120, 895), venue_font, NAVY_DARK)


# ============================================================
# TABLE DRAWING
# ============================================================

def draw_table_header(draw, x1, x2, y, title, color):
    h = 56

    draw.rounded_rectangle(
        (x1, y, x2, y + h),
        radius=14,
        fill=color,
        outline=GOLD,
        width=2,
    )

    draw_center(
        draw,
        title,
        (x1, y, x2, y + h),
        get_font(28, bold=True),
        WHITE,
    )

    return y + h


def draw_main_prizes(draw):
    x1 = 80
    x2 = WIDTH - 80
    y = 925

    table_top = draw_table_header(
        draw,
        x1,
        x2,
        y,
        "MAIN PRIZES",
        MAROON,
    )

    header_h = 48
    row_h = 48

    prize_col = 320

    # Column header (dark maroon, gold text).
    draw.rectangle(
        (x1, table_top, x1 + prize_col, table_top + header_h),
        fill=MAROON_DARK,
        outline=GOLD,
        width=2,
    )

    draw.rectangle(
        (x1 + prize_col, table_top, x2, table_top + header_h),
        fill=MAROON_DARK,
        outline=GOLD,
        width=2,
    )

    draw_center(
        draw,
        "PRIZE",
        (
            x1,
            table_top,
            x1 + prize_col,
            table_top + header_h,
        ),
        get_font(20, bold=True),
        GOLD_LIGHT,
    )

    draw_center(
        draw,
        "GIFT ITEM",
        (
            x1 + prize_col,
            table_top,
            x2,
            table_top + header_h,
        ),
        get_font(20, bold=True),
        GOLD_LIGHT,
    )

    for i, (prize, gift) in enumerate(MAIN_PRIZES):
        top = table_top + header_h + i * row_h
        bottom = top + row_h

        row_fill = CREAM_2 if i % 2 == 0 else GOLD_PALE

        draw.rectangle(
            (x1, top, x1 + prize_col, bottom),
            fill=row_fill,
            outline=GOLD_LINE,
            width=1,
        )

        draw.rectangle(
            (x1 + prize_col, top, x2, bottom),
            fill=row_fill,
            outline=GOLD_LINE,
            width=1,
        )

        draw_center(
            draw,
            prize,
            (
                x1,
                top,
                x1 + prize_col,
                bottom,
            ),
            get_font(20, bold=True),
            MAROON_DARK,
        )

        gift_font = fit_font(
            draw,
            gift,
            (x2 - x1 - prize_col) - 15,
            22,
            bold=False,
        )

        draw_mixed_left(
            draw,
            gift,
            (
                x1 + prize_col + 5,
                top,
                x2 - 5,
                bottom,
            ),
            gift_font,
            get_emoji_font(gift_font.size),
            BLACK,
        )

    # Prominent outer border around the whole table.
    table_bottom = table_top + header_h + len(MAIN_PRIZES) * row_h
    draw.rounded_rectangle(
        (x1 - 4, y - 4, x2 + 4, table_bottom + 4),
        radius=10,
        outline=GOLD,
        width=2,
    )


def draw_consolation_prizes(draw):
    x1 = 80
    x2 = WIDTH - 80
    y = 1525

    table_top = draw_table_header(
        draw,
        x1,
        x2,
        y,
        "CONSOLATION PRIZES",
        NAVY,
    )

    header_h = 48
    row_h = 48

    category_col = 160
    qty_col = 100

    # Column headers (dark navy, gold text).
    draw.rectangle(
        (x1, table_top, x1 + category_col, table_top + header_h),
        fill=NAVY_DARK,
        outline=GOLD,
        width=2,
    )

    draw.rectangle(
        (x1 + category_col, table_top, x2 - qty_col, table_top + header_h),
        fill=NAVY_DARK,
        outline=GOLD,
        width=2,
    )

    draw.rectangle(
        (x2 - qty_col, table_top, x2, table_top + header_h),
        fill=NAVY_DARK,
        outline=GOLD,
        width=2,
    )

    draw_center(
        draw,
        "CATEGORY",
        (
            x1,
            table_top,
            x1 + category_col,
            table_top + header_h,
        ),
        get_font(20, bold=True),
        GOLD_LIGHT,
    )

    draw_center(
        draw,
        "GIFT ITEM",
        (
            x1 + category_col,
            table_top,
            x2 - qty_col,
            table_top + header_h,
        ),
        get_font(20, bold=True),
        GOLD_LIGHT,
    )

    draw_center(
        draw,
        "QTY",
        (
            x2 - qty_col,
            table_top,
            x2,
            table_top + header_h,
        ),
        get_font(20, bold=True),
        GOLD_LIGHT,
    )

    for i, (category, gift, qty) in enumerate(CONSOLATION_PRIZES):
        top = table_top + header_h + i * row_h
        bottom = top + row_h

        row_fill = CREAM_2 if i % 2 == 0 else GOLD_PALE

        draw.rectangle(
            (x1, top, x1 + category_col, bottom),
            fill=row_fill,
            outline=GOLD_LINE,
            width=1,
        )

        draw.rectangle(
            (x1 + category_col, top, x2 - qty_col, bottom),
            fill=row_fill,
            outline=GOLD_LINE,
            width=1,
        )

        draw.rectangle(
            (x2 - qty_col, top, x2, bottom),
            fill=row_fill,
            outline=GOLD_LINE,
            width=1,
        )

        draw_center(
            draw,
            category,
            (
                x1,
                top,
                x1 + category_col,
                bottom,
            ),
            get_font(20, bold=True),
            MAROON_DARK,
        )

        gift_font = fit_font(
            draw,
            gift,
            (x2 - qty_col - x1 - category_col) - 15,
            22,
            bold=False,
        )

        draw_mixed_left(
            draw,
            gift,
            (
                x1 + category_col + 5,
                top,
                x2 - qty_col - 5,
                bottom,
            ),
            gift_font,
            get_emoji_font(gift_font.size),
            BLACK,
        )

        draw_center(
            draw,
            qty,
            (
                x2 - qty_col,
                top,
                x2,
                bottom,
            ),
            get_font(20, bold=True),
            MAROON_DARK,
        )

    # Prominent outer border around the whole table.
    table_bottom = table_top + header_h + len(CONSOLATION_PRIZES) * row_h
    draw.rounded_rectangle(
        (x1 - 4, y - 4, x2 + 4, table_bottom + 4),
        radius=10,
        outline=GOLD,
        width=2,
    )

    # Thank-you card fills the unused area underneath.
    card_top = table_top + header_h + (4 * row_h) + 25
    card_bottom = 1955

    # Card background.
    draw.rounded_rectangle(
        (
            x1,
            card_top,
            x2,
            card_bottom,
        ),
        radius=18,
        fill=WHITE,
        outline=GOLD,
        width=3,
    )

    # Gold corner accents on the card.
    for cx, cy, sx, sy in [
        (x1 + 14, card_top + 14, 1, 1),
        (x2 - 14, card_top + 14, -1, 1),
        (x1 + 14, card_bottom - 14, 1, -1),
        (x2 - 14, card_bottom - 14, -1, -1),
    ]:
        d = 5
        draw.polygon(
            [
                (cx, cy - d),
                (cx + d, cy),
                (cx, cy + d),
                (cx - d, cy),
            ],
            fill=GOLD,
        )

    # Maroon header band inside the card.
    band_top = card_top + 12
    band_bottom = card_top + 58

    draw.rounded_rectangle(
        (
            x1 + 8,
            band_top,
            x2 - 8,
            band_bottom,
        ),
        radius=12,
        fill=MAROON,
        outline=GOLD,
        width=2,
    )

    # Gold heart icon on the left of the band.
    heart_cx = x1 + 60
    heart_cy = (band_top + band_bottom) // 2
    hs = 10
    draw.polygon(
        [
            (heart_cx, heart_cy + hs),
            (heart_cx - hs, heart_cy),
            (heart_cx - hs // 2, heart_cy - hs),
            (heart_cx, heart_cy - hs // 2),
            (heart_cx + hs // 2, heart_cy - hs),
            (heart_cx + hs, heart_cy),
        ],
        fill=GOLD_LIGHT,
    )

    # Gold heart icon on the right of the band.
    heart_cx = x2 - 60
    draw.polygon(
        [
            (heart_cx, heart_cy + hs),
            (heart_cx - hs, heart_cy),
            (heart_cx - hs // 2, heart_cy - hs),
            (heart_cx, heart_cy - hs // 2),
            (heart_cx + hs // 2, heart_cy - hs),
            (heart_cx + hs, heart_cy),
        ],
        fill=GOLD_LIGHT,
    )

    draw_center(
        draw,
        "Thank You For Your Support",
        (
            x1 + 8,
            band_top,
            x2 - 8,
            band_bottom,
        ),
        get_font(30, bold=True, italic=True),
        GOLD_LIGHT,
    )

    # Team message — positioned just below the band.
    msg_text = "TEAM JAIBHADRA FOUNDATION"

    avail_w = x2 - x1 - 20
    msg_font = fit_font(draw, msg_text, avail_w, 34, bold=True)

    # Positioned centered in the space below the band.
    line_h = int(msg_font.size * 1.4)
    remaining_top = band_bottom
    remaining_bottom = card_bottom
    start_y = remaining_top + ((remaining_bottom - remaining_top) - line_h) // 2

    draw_center(
        draw,
        msg_text,
        (
            x1 + 10,
            start_y,
            x2 - 10,
            start_y + line_h,
        ),
        msg_font,
        MAROON_DARK,
    )


# ============================================================
# FOOTER
# ============================================================

def draw_footer(draw, start, end, sample=False):
    is_range = start != end
    serial_text = f"{start:04d}-{end:04d}" if is_range else f"{start:04d}"

    # Contact us bar (10px gap below thank-you card which ends at 1955).
    contact_bar = (100, 1965, WIDTH - 100, 2025)
    draw.rounded_rectangle(contact_bar, radius=20, fill=CREAM_2, outline=GOLD, width=2)

    for cx in (100, WIDTH - 100):
        d = 5
        draw.polygon([(cx, 1995 - d), (cx + d, 1995), (cx, 1995 + d), (cx - d, 1995)], fill=GOLD)

    # Phone icon
    ph_cx = 140
    ph_cy = 1995
    draw.rounded_rectangle((ph_cx - 8, ph_cy - 10, ph_cx + 8, ph_cy + 10), radius=4, outline=MAROON_DARK, width=2)
    draw.line((ph_cx - 4, ph_cy - 5, ph_cx + 4, ph_cy - 5), fill=MAROON_DARK, width=2)
    draw.ellipse((ph_cx - 2, ph_cy + 3, ph_cx + 2, ph_cy + 7), fill=MAROON_DARK)

    draw_center(draw, "CONTACT US  :  9810117382  |  9811398177  |  9818384890",
                (170, 1965, WIDTH - 100, 2025),
                get_font(22, bold=True), MAROON_DARK)

    # Footer bar (10px gap below contact bar).
    footer = (45, 2035, WIDTH - 45, 2105)
    draw.rounded_rectangle(footer, radius=18, fill=MAROON, outline=GOLD, width=3)

    # Bottom line: SERIAL (left) | FOR DETAILS VISIT... (center) | GOOD LUCK (right).
    serial_label = f"SAMPLE {serial_text}" if sample else f"SERIAL {serial_text}"
    draw_center(draw, serial_label, (60, 2041, 360, 2100),
                get_font(16, bold=True), GOLD_LIGHT)

    draw_center(draw, "FOR DETAILS VISIT WWW.JAIBHADRA.ORG",
                (360, 2041, WIDTH - 360, 2100),
                get_font(20, bold=True), WHITE)

    draw_center(draw, "GOOD LUCK", (WIDTH - 360, 2041, WIDTH - 60, 2100),
                get_font(16, bold=True), GOLD_LIGHT)


# ============================================================
# ONE COUPON
# ============================================================
def _render_coupon(start, end, buyer=None, phone=None, address=None, amount=None, sample=False):
    """Build and return the coupon PIL.Image in memory.  Does NOT save
    anything to disk — used by the live preview so we don't litter the
    output folder with a PNG on every keystroke."""
    image = Image.new(
        "RGB",
        (WIDTH, HEIGHT),
        CREAM,
    )

    draw = ImageDraw.Draw(image)

    draw_ticket_border(draw)
    draw_text_watermark(draw)

    # Photos flank the header on left and right sides.
    photo_w = 200
    photo_h = 200
    photo_y = 60
    draw_photo_card(draw, image, LEFT_PHOTO, (60, photo_y, 60 + photo_w, photo_y + photo_h))
    draw_photo_card(draw, image, RIGHT_PHOTO, (WIDTH - 60 - photo_w, photo_y, WIDTH - 60, photo_y + photo_h))

    draw_header(draw)

    draw_coupon_center(
        draw,
        image,
        start,
        end,
        buyer=buyer,
        phone=phone,
        address=address,
        amount=amount,
        sample=sample,
    )

    draw_date_box(draw)

    draw_main_prizes(draw)

    draw_consolation_prizes(draw)

    draw_footer(draw, start, end, sample=sample)

    return image


def create_coupon(start, end, buyer=None, phone=None, address=None, amount=None):
    image = _render_coupon(start, end, buyer=buyer, phone=phone, address=address, amount=amount)

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    is_range = start != end
    num_part = f"{start:04d}-{end:04d}" if is_range else f"{start:04d}"

    # Build a filesystem-safe filename including the buyer's name.
    safe_buyer = ""
    if buyer:
        invalid = '<>:"/\\|?*'
        safe_buyer = "".join("_" if ch in invalid else ch for ch in buyer).strip()
        safe_buyer = safe_buyer.replace(" ", "_")

    if safe_buyer:
        filename = f"{safe_buyer}_{num_part}.png"
    else:
        filename = f"coupon_{num_part}.png"

    output_path = OUTPUT_DIR / filename

    # PNG keeps the design crisp.
    image.save(
        output_path,
        "PNG",
        optimize=True,
        dpi=(300, 300),
    )

    return output_path.name


# ============================================================
# EXCEL TRACKING
# ============================================================

SALES_HEADERS = ["S.No", "Buyer Name", "Phone", "Address", "Start No", "End No", "Qty", "Date Sold", "Type", "Set Size", "Donation Amount"]

# 1-based column indices for the fixed columns (used by the delete helpers
# so they don't break when new columns are appended to SALES_HEADERS).
COL_SNO = 1
COL_NAME = 2
COL_PHONE = 3
COL_ADDRESS = 4
COL_START = 5
COL_END = 6
COL_QTY = 7
COL_DATE = 8
COL_TYPE = 9
COL_SET_SIZE = 10
COL_AMOUNT = 11


def _row_set_size(start, end, stype):
    """Compute the Set Size value for a backfilled row during migration.
    For PHYSICAL rows the set size is the strip length (end-start+1).
    For normal SALE rows there is no single 'set' concept, so it is blank."""
    if stype == "PHYSICAL":
        try:
            return int(end) - int(start) + 1
        except (TypeError, ValueError):
            return ""
    return ""


def _row_donation(start, end, stype):
    """Compute the Donation Amount value for a backfilled row during
    migration.  A physical set's donation = SET_PRICES lookup (or
    qty * 100 fallback).  A normal sale = qty * 100."""
    try:
        qty = int(end) - int(start) + 1
    except (TypeError, ValueError):
        return ""
    if stype == "PHYSICAL":
        return price_for_set_size(qty)
    return qty * PRICE_PER_COUPON


def ensure_sales_file():
    """Create coupon_sales.xlsx with headers if it does not exist.
    Migrates older files (without the 'Type' column) by adding it and
    backfilling existing rows with 'SALE'.
    Raises RuntimeError if the file exists but is locked.

    Fast path: once we have successfully opened the file once in this
    process, subsequent calls just confirm the file still exists and
    return immediately — we do NOT re-open the workbook every call (that
    was a major source of UI lag, since several code paths call this on
    every refresh)."""
    global _ensure_done

    if not SALES_FILE.exists():
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Coupon Sales"
        ws.append(SALES_HEADERS)

        # Bold header row.
        for col_idx in range(1, len(SALES_HEADERS) + 1):
            ws.cell(row=1, column=col_idx).font = openpyxl.styles.Font(bold=True)

        wb.save(SALES_FILE)
        wb.close()
        _ensure_done = True
        return

    # Fast path: we validated the file earlier in this session and the
    # cache hasn't been invalidated (no write since).  Skip the re-open.
    if _ensure_done:
        return

    # File exists — probe that it is not corrupt by trying to open it.
    # If it is locked (open in Excel / OneDrive), we do NOT raise here:
    # reading can still proceed via the in-memory cache / a read-only open,
    # and only a *write* (record_sale, delete_*) needs to fail with a clear
    # message.  Raising here made the whole app unusable whenever Excel was
    # open on the file.
    try:
        wb = openpyxl.load_workbook(SALES_FILE)
    except PermissionError:
        # File is locked.  Skip the probe — mark done so we don't retry on
        # every call.  Reads use the cache; writes will raise their own
        # PermissionError-derived RuntimeError.
        _ensure_done = True
        return
    except (zipfile.BadZipFile, KeyError, ValueError, EOFError):
        # The file exists but is corrupted (e.g. truncated by a crash/OneDrive
        # sync issue).  Preserve the broken copy as a backup and recreate a
        # fresh workbook with headers so the app can still start.  Existing
        # sales data in the corrupt file is already unreadable.
        backup = SALES_FILE.with_name(
            f"coupon_sales_corrupt_{datetime.now():%Y%m%d_%H%M%S}.bak"
        )
        try:
            shutil.copy2(SALES_FILE, backup)
        except Exception:
            pass
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Coupon Sales"
        ws.append(SALES_HEADERS)
        for col_idx in range(1, len(SALES_HEADERS) + 1):
            ws.cell(row=1, column=col_idx).font = openpyxl.styles.Font(bold=True)
        try:
            wb.save(SALES_FILE)
        except PermissionError:
            raise RuntimeError(
                f"Cannot save '{SALES_FILE.name}'. Please CLOSE it in "
                f"Microsoft Excel / OneDrive and try again."
            )
        wb.close()
        _ensure_done = True
        _invalidate_sales_cache()
        return

    ws = wb.active

    # Migration: add 'Type' column to older files that lack it.
    header_cells = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    header_list = list(header_cells)

    if header_list[-1] != SALES_HEADERS[-1]:
        # The file predates the current schema (missing 'Set Size' and/or
        # 'Donation Amount', or even missing 'Type').  Write the full header
        # set and backfill every existing row.
        for col_idx, header in enumerate(SALES_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True)

        # Backfill existing data rows.
        for row_idx in range(2, ws.max_row + 1):
            start_val = ws.cell(row=row_idx, column=5).value
            end_val = ws.cell(row=row_idx, column=6).value
            # Recover the Type column (was missing in the very first schema).
            existing_type = ws.cell(row=row_idx, column=9).value
            if existing_type is None:
                any_data = any(
                    ws.cell(row=row_idx, column=c).value is not None
                    for c in range(1, 9)
                )
                if any_data:
                    existing_type = "SALE"
                    ws.cell(row=row_idx, column=9).value = "SALE"
                else:
                    existing_type = ""
            stype = existing_type or "SALE"

            # Set Size (column 10): strip length for PHYSICAL, blank for SALE.
            set_sz = _row_set_size(start_val, end_val, stype)
            ws.cell(row=row_idx, column=10).value = set_sz

            # Donation Amount (column 11).
            ws.cell(row=row_idx, column=11).value = _row_donation(
                start_val, end_val, stype
            )

        try:
            wb.save(SALES_FILE)
        except PermissionError:
            raise RuntimeError(
                f"Cannot save '{SALES_FILE.name}'. Please CLOSE it in "
                f"Microsoft Excel / OneDrive and try again."
            )
        finally:
            wb.close()
        _ensure_done = True
        _invalidate_sales_cache()
        return

    wb.close()
    _ensure_done = True


def _load_sales_workbook(read_only=False):
    """Load coupon_sales.xlsx. Raises RuntimeError with a friendly message
    if the file is locked (e.g. open in Excel).  Pass read_only=True for
    read paths — when the file is locked we transparently read from a
    temporary copy so the UI stays responsive even while Excel is open."""
    try:
        return openpyxl.load_workbook(SALES_FILE, read_only=read_only)
    except PermissionError:
        if not read_only:
            # Try a read-only open before giving up — sometimes succeeds
            # even when Excel has the file open.
            try:
                return openpyxl.load_workbook(SALES_FILE, read_only=True)
            except PermissionError:
                raise RuntimeError(
                    f"Cannot open '{SALES_FILE.name}'. Please CLOSE it in "
                    f"Microsoft Excel / OneDrive and try again."
                )
        # Read path: the file is locked (Excel/OneDrive has an exclusive
        # lock).  Copy it to a temp file and read from there — the data is
        # still perfectly readable, we just can't write back until the lock
        # is released.
        try:
            import tempfile
            tmp = tempfile.NamedTemporaryFile(
                suffix=".xlsx", delete=False, prefix="coupon_sales_read_"
            )
            tmp.close()
            shutil.copy2(SALES_FILE, tmp.name)
            wb = openpyxl.load_workbook(tmp.name, read_only=True)
            # Tag the workbook so we clean up the temp file on close().
            wb._jbf_tmp_path = tmp.name
            return wb
        except Exception as exc:
            raise RuntimeError(
                f"Cannot open '{SALES_FILE.name}'. Please CLOSE it in "
                f"Microsoft Excel / OneDrive and try again. ({exc})"
            )
    except (zipfile.BadZipFile, KeyError, ValueError, EOFError) as exc:
        raise RuntimeError(
            f"'{SALES_FILE.name}' is corrupted (not a valid xlsx). "
            f"Delete it or restore from backup and try again. ({exc})"
        )


def get_sold_ranges():
    """Return a list of (start, end) tuples already sold, read from the
    cached sales rows (reloaded from Excel only when the file changed)."""
    if not SALES_FILE.exists():
        return []

    ranges = []
    for row in _get_sales_rows():
        # Columns: S.No, Name, Phone, Address, Start, End, Qty, Date, Type
        if len(row) < 6:
            continue
        start = row[4]
        end = row[5]
        if start is None or end is None:
            continue
        try:
            ranges.append((int(start), int(end)))
        except (ValueError, TypeError):
            continue

    return ranges


def is_already_sold(start, end):
    """Return the (start, end) of an overlapping sold range, or None if free.
    Raises RuntimeError if the Excel file is locked."""
    for s, e in get_sold_ranges():
        if start <= e and end >= s:
            return (s, e)
    return None


def record_sale(name, phone, address, start, end, sale_type="SALE",
                 set_size=None, amount=None):
    """Append a sale row to coupon_sales.xlsx.

    For physical sets, pass sale_type='PHYSICAL' and leave name/phone/address
    as None — those buyer fields are written blank so the row locks the
    coupon numbers without any buyer details.

    set_size is the strip length (for PHYSICAL rows); pass None for normal
    sales.  amount is the total donation for this row; when None it is
    computed from the sale type (qty * 100 for a sale, SET_PRICES for a
    physical set).

    Raises RuntimeError if the Excel file is locked."""
    ensure_sales_file()

    wb = _load_sales_workbook()
    ws = wb.active

    next_sno = ws.max_row  # header is row 1, so max_row gives next S.No
    if next_sno < 1:
        next_sno = 1

    qty = end - start + 1
    date_sold = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Resolve the set_size + amount when the caller omitted them so that
    # older callers (and the migration path) still produce sensible rows.
    if set_size is None:
        set_size = qty if sale_type == "PHYSICAL" else ""
    if amount is None:
        if sale_type == "PHYSICAL":
            amount = price_for_set_size(qty)
        else:
            amount = qty * PRICE_PER_COUPON

    ws.append([
        next_sno, name, phone, address, start, end, qty, date_sold,
        sale_type, set_size, amount,
    ])

    try:
        wb.save(SALES_FILE)
    except PermissionError:
        raise RuntimeError(
            f"Cannot save '{SALES_FILE.name}'. Please CLOSE it in "
            f"Microsoft Excel / OneDrive and try again."
        )
    finally:
        wb.close()
        _invalidate_sales_cache()


def list_sales(print_it=True):
    """Print all sales rows as a formatted table.
    Returns the list of rows (excluding header). Raises RuntimeError if locked.

    Uses the in-memory cache so repeated UI refreshes do not re-open Excel."""
    rows = _get_sales_rows()

    if print_it:
        print()
        print("-" * 124)
        print(f"{'S.No':<6}{'Name':<20}{'Phone':<14}{'Start':<8}{'End':<8}{'Qty':<6}{'Type':<10}{'Donation':<12}{'Date Sold'}")
        print("-" * 124)

        if not rows:
            print("(no sales recorded yet)")
        else:
            for r in rows:
                # Tolerate rows that predate the new columns.
                if len(r) >= 11:
                    sno, name, phone, address, start, end, qty, date, stype, set_sz, amount = r[:11]
                elif len(r) >= 9:
                    sno, name, phone, address, start, end, qty, date, stype = r[:9]
                    amount = qty * PRICE_PER_COUPON if stype != "PHYSICAL" else price_for_set_size(qty)
                else:
                    sno, name, phone, address, start, end, qty, date = r
                    stype = "SALE"
                    amount = qty * PRICE_PER_COUPON
                name_str = (str(name)[:18] if name else "<PHYSICAL>") if stype == "PHYSICAL" else (str(name)[:18] if name else "")
                phone_str = str(phone)[:12] if phone else ""
                stype_str = str(stype)[:8] if stype else "SALE"
                amt_str = f"\u20B9{amount}" if amount is not None else ""
                print(f"{sno!s:<6}{name_str:<20}{phone_str:<14}{start!s:<8}{end!s:<8}{qty!s:<6}{stype_str:<10}{amt_str:<12}{date}")

        print("-" * 106)
        print()

    return rows


def delete_sale(sno, delete_png=False):
    """Delete the sale row with the given S.No and renumber remaining rows.
    If delete_png is True, also remove the corresponding PNG file.
    Returns True if a row was deleted, False if S.No was not found.
    Raises RuntimeError if the Excel file is locked."""
    ensure_sales_file()
    wb = _load_sales_workbook()
    ws = wb.active

    target_row = None
    target_range = None  # (start, end) of the row being deleted
    target_buyer = None

    for row_idx in range(2, ws.max_row + 1):
        cell_sno = ws.cell(row=row_idx, column=1).value
        if cell_sno is None:
            continue
        try:
            if int(cell_sno) == int(sno):
                target_row = row_idx
                start = ws.cell(row=row_idx, column=5).value
                end = ws.cell(row=row_idx, column=6).value
                target_range = (int(start), int(end))
                target_buyer = ws.cell(row=row_idx, column=2).value
                break
        except (ValueError, TypeError):
            continue

    if target_row is None:
        return False

    ws.delete_rows(target_row, 1)

    # Renumber S.No for all remaining rows so they stay sequential.
    new_sno = 1
    for row_idx in range(2, ws.max_row + 1):
        cell_sno = ws.cell(row=row_idx, column=1).value
        if cell_sno is None:
            continue
        ws.cell(row=row_idx, column=1).value = new_sno
        new_sno += 1

    try:
        wb.save(SALES_FILE)
    except PermissionError:
        raise RuntimeError(
            f"Cannot save '{SALES_FILE.name}'. Please CLOSE it in "
            f"Microsoft Excel / OneDrive and try again."
        )
    finally:
        wb.close()
        _invalidate_sales_cache()

    # Optionally delete the PNG file.
    if delete_png and target_range is not None:
        start, end = target_range
        is_range = start != end
        num_part = f"{start:04d}-{end:04d}" if is_range else f"{start:04d}"

        # Build the same filesystem-safe filename create_coupon would have used.
        safe_buyer = ""
        if target_buyer:
            invalid = '<>:"/\\|?*'
            safe_buyer = "".join("_" if ch in invalid else ch for ch in str(target_buyer)).strip()
            safe_buyer = safe_buyer.replace(" ", "_")

        filename = f"{safe_buyer}_{num_part}.png" if safe_buyer else f"coupon_{num_part}.png"
        png_path = OUTPUT_DIR / filename

        if png_path.exists():
            try:
                png_path.unlink()
                print(f"Deleted PNG: {png_path.name}")
            except Exception as exc:
                print(f"WARNING: Could not delete {png_path.name}: {exc}")
        else:
            print(f"PNG not found (already removed): {png_path.name}")

    return True


# ============================================================
# INTERACTIVE SALE
# ============================================================

def ask(prompt):
    """input() wrapper that strips whitespace and returns None on empty/quit."""
    raw = input(prompt).strip()
    if raw.lower() in ("q", "quit", "exit"):
        return None
    return raw


def ask_int(prompt):
    """Prompt for a non-negative integer. Returns None on quit."""
    while True:
        raw = ask(prompt)
        if raw is None:
            return None
        try:
            value = int(raw)
            if value < 0:
                print("  Number cannot be negative. Try again.")
                continue
            return value
        except ValueError:
            print("  Please enter a valid number.")


def prompt_sale():
    """Ask the user for one sale's details, validate, generate coupons, record sale.
    Returns False to stop the loop, True to continue."""
    print()
    print("-" * 72)
    print("NEW SALE")
    print("(type 'q' at any prompt to finish)")
    print("-" * 72)

    name = ask("Buyer Name : ")
    if name is None:
        return False

    phone = ask("Phone      : ")
    if phone is None:
        return False

    address = ask("Address    : ")
    if address is None:
        return False

    start = ask_int("Start Number: ")
    if start is None:
        return False

    if start > MAX_COUPON:
        print(f"ERROR: Start number cannot exceed {MAX_COUPON}. Sale cancelled.")
        return True

    qty = ask_int("Quantity    : ")
    if qty is None:
        return False

    if qty < 1:
        print("ERROR: Quantity must be at least 1. Sale cancelled.")
        return True

    # Auto-calculate end number from start + quantity.
    end = start + qty - 1

    if end > MAX_COUPON:
        print(f"ERROR: End number ({end:04d}) would exceed {MAX_COUPON}. Sale cancelled.")
        return True

    print(f"  -> End Number: {end:04d}")

    try:
        overlap = is_already_sold(start, end)
    except RuntimeError as exc:
        print(f"ERROR: {exc}")
        print("       Sale cancelled.")
        return True

    if overlap:
        print(f"ERROR: Coupons {overlap[0]:04d}-{overlap[1]:04d} are already sold.")
        print("       Sale cancelled to prevent double-selling.")
        return True

    # Generate a single PNG for the sale (single coupon or whole range).
    print()
    qty = end - start + 1
    total_amount = qty * PRICE_PER_COUPON
    if start == end:
        print(f"Generating 1 coupon: {start:04d}")
    else:
        print(f"Generating {qty} coupons: {start:04d} - {end:04d}")
    print(f"  Donation amount: \u20B9{total_amount}")

    filename = create_coupon(
        start, end, buyer=name, phone=phone, address=address, amount=total_amount
    )
    print(f"  Saved {filename}")

    # Record sale in Excel.
    try:
        record_sale(
            name, phone, address, start, end, amount=total_amount
        )
        print(f"Recorded sale in {SALES_FILE.name}")
    except RuntimeError as exc:
        print(f"WARNING: {exc}")
        print("         Coupons were generated but the sale was NOT recorded.")

    print()
    print("Sale complete.")

    return True


def prompt_delete():
    """Interactive flow to delete a sale. Returns True to continue the menu loop."""
    try:
        rows = list_sales()
    except RuntimeError as exc:
        print(f"ERROR: {exc}")
        return True

    if not rows:
        print("There is nothing to delete.")
        return True

    sno_raw = ask("Enter S.No to delete (or 'q' to cancel): ")
    if sno_raw is None:
        print("Delete cancelled.")
        return True

    try:
        sno = int(sno_raw)
    except ValueError:
        print("Invalid S.No. Delete cancelled.")
        return True

    # Confirm.
    confirm = ask(f"Are you sure you want to delete sale #{sno}? (y/n): ")
    if confirm is None or confirm.lower() not in ("y", "yes"):
        print("Delete cancelled.")
        return True

    # Ask whether to also delete the PNG.
    del_png = False
    png_ans = ask("Also delete the coupon PNG file? (y/n): ")
    if png_ans is not None and png_ans.lower() in ("y", "yes"):
        del_png = True

    try:
        deleted = delete_sale(sno, delete_png=del_png)
    except RuntimeError as exc:
        print(f"ERROR: {exc}")
        return True

    if deleted:
        print(f"Sale #{sno} deleted successfully.")
        if del_png:
            print("Coupon numbers from this sale are now free for re-sale.")
        else:
            print("PNG file kept. Coupon numbers are now free for re-sale.")
    else:
        print(f"No sale found with S.No #{sno}.")

    return True


def delete_all_sales(delete_pngs=False):
    """Delete every sale row from the Excel file (keep the header).
    If delete_pngs is True, also delete all generated PNG files.
    Returns the number of rows deleted.
    Raises RuntimeError if the Excel file is locked."""
    ensure_sales_file()
    wb = _load_sales_workbook()
    ws = wb.active

    # Count existing sale rows (row 1 is the header).
    count = max(0, ws.max_row - 1)

    if count == 0:
        return 0

    # Remove all data rows, keep only the header.
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    try:
        wb.save(SALES_FILE)
    except PermissionError:
        raise RuntimeError(
            f"Cannot save '{SALES_FILE.name}'. Please CLOSE it in "
            f"Microsoft Excel / OneDrive and try again."
        )
    finally:
        wb.close()
        _invalidate_sales_cache()

    # Optionally delete every PNG in the output folder.
    if delete_pngs and OUTPUT_DIR.exists():
        removed = 0
        for png in OUTPUT_DIR.glob("*.png"):
            try:
                png.unlink()
                removed += 1
            except Exception:
                pass
        if removed:
            print(f"Deleted {removed} PNG file(s) from {OUTPUT_DIR.name}.")

    return count


def prompt_delete_all():
    """Interactive flow to delete ALL sales at once.
    Returns True to continue the menu loop."""
    try:
        rows = list_sales()
    except RuntimeError as exc:
        print(f"ERROR: {exc}")
        return True

    if not rows:
        print("There are no sales to delete.")
        return True

    print(f"This will delete ALL {len(rows)} sale(s) from the Excel tracker.")
    confirm = ask("Are you sure? This cannot be undone. (y/n): ")
    if confirm is None or confirm.lower() not in ("y", "yes"):
        print("Delete-all cancelled.")
        return True

    del_png = False
    png_ans = ask("Also delete ALL generated PNG files? (y/n): ")
    if png_ans is not None and png_ans.lower() in ("y", "yes"):
        del_png = True

    try:
        count = delete_all_sales(delete_pngs=del_png)
    except RuntimeError as exc:
        print(f"ERROR: {exc}")
        return True

    print(f"Deleted all {count} sale(s). Excel tracker is now empty.")
    if del_png:
        print("All coupon numbers are free for re-sale.")
    else:
        print("PNG files kept. All coupon numbers are free for re-sale.")

    return True


def show_last_sale():
    """Print the most recent sale. Returns True to continue the menu loop."""
    try:
        rows = list_sales(print_it=False)
    except RuntimeError as exc:
        print(f"ERROR: {exc}")
        return True

    if not rows:
        print()
        print("No sales recorded yet.")
        return True

    last = rows[-1]
    if len(last) >= 11:
        sno, name, phone, address, start, end, qty, date, stype, set_sz, amount = last[:11]
    elif len(last) >= 9:
        sno, name, phone, address, start, end, qty, date, stype = last[:9]
        amount = qty * PRICE_PER_COUPON if stype != "PHYSICAL" else price_for_set_size(qty)
    else:
        sno, name, phone, address, start, end, qty, date = last
        stype = "SALE"
        amount = qty * PRICE_PER_COUPON

    print()
    print("=" * 50)
    print("LAST COUPON SOLD")
    print("=" * 50)
    print(f"  S.No     : {sno}")
    print(f"  Type     : {stype}")
    print(f"  Name     : {name if name else '<PHYSICAL>'}")
    print(f"  Phone    : {phone if phone else '-'}")
    print(f"  Address  : {address if address else '-'}")
    print(f"  Start    : {start:04d}" if isinstance(start, int) else f"  Start    : {start}")
    print(f"  End      : {end:04d}" if isinstance(end, int) else f"  End      : {end}")
    print(f"  Quantity : {qty}")
    print(f"  Donation : \u20B9{amount}" if amount is not None else "")
    print(f"  Date     : {date}")
    print("=" * 50)

    return True


def prompt_physical_set():
    """Generate one or more physical-sale coupon sets with NO buyer details.

    A set is a strip of `set_size` consecutive coupon numbers (10, 5, or 1).
    Each set becomes one PNG (a range strip) and one row in the Excel tracker
    with Type='PHYSICAL' and blank buyer fields, so those numbers are locked
    exactly like a normal sale but carry no buyer information.

    The QR on each generated coupon shows only the coupon number(s) — no
    buyer name/phone/address, because create_coupon() omits those lines when
    buyer/phone/address are None.

    Returns True to continue the menu loop, False to stop."""
    print()
    print("-" * 72)
    print("PHYSICAL SET GENERATION")
    print("(no buyer details — numbers are locked for physical sale)")
    print("(type 'q' at any prompt to cancel)")
    print("-" * 72)

    print()
    print("Choose set size:")
    print(f"  1. Set of 10  (one strip covers 10 coupon numbers)  \u20B9{price_for_set_size(10)}")
    print(f"  2. Set of 5   (one strip covers 5 coupon numbers)  \u20B9{price_for_set_size(5)}")
    print(f"  3. Set of 1   (one coupon per set)                 \u20B9{price_for_set_size(1)}")
    size_choice = ask("Choice (1/2/3): ")
    if size_choice is None:
        return True

    size_map = {"1": 10, "2": 5, "3": 1}
    if size_choice not in size_map:
        print("Invalid set-size choice. Cancelled.")
        return True
    set_size = size_map[size_choice]

    start = ask_int("Start Number : ")
    if start is None:
        return True
    if start < 1:
        print("ERROR: Start number must be at least 1. Cancelled.")
        return True
    if start > MAX_COUPON:
        print(f"ERROR: Start number cannot exceed {MAX_COUPON}. Cancelled.")
        return True

    num_sets = ask_int(f"Number of sets of {set_size}: ")
    if num_sets is None:
        return True
    if num_sets < 1:
        print("ERROR: You must generate at least one set. Cancelled.")
        return True

    total_coupons = num_sets * set_size
    end = start + total_coupons - 1

    if end > MAX_COUPON:
        print(
            f"ERROR: This would end at {end:04d}, which exceeds the max of "
            f"{MAX_COUPON}. Cancelled."
        )
        return True

    set_price = price_for_set_size(set_size)
    total_amount = num_sets * set_price

    print()
    print(
        f"Plan: {num_sets} set(s) of {set_size} = {total_coupons} coupons, "
        f"numbering {start:04d} - {end:04d}."
    )
    print(
        f"Donation: \u20B9{set_price} per set x {num_sets} set(s) = \u20B9{total_amount} total."
    )
    print("Each set is one PNG strip, no buyer details, QR shows only the coupon number.")
    confirm = ask("Proceed? (y/n): ")
    if confirm is None or confirm.lower() not in ("y", "yes"):
        print("Cancelled. No coupons generated.")
        return True

    # Pre-check overlap for the WHOLE block so we don't generate half a batch.
    try:
        overlap = is_already_sold(start, end)
    except RuntimeError as exc:
        print(f"ERROR: {exc}")
        print("       Cancelled.")
        return True

    if overlap:
        print(
            f"ERROR: Coupons {overlap[0]:04d}-{overlap[1]:04d} are already "
            f"assigned (sold or physical). Cancelled to prevent double-booking."
        )
        return True

    print()
    generated = 0
    failed = 0

    for i in range(num_sets):
        s = start + i * set_size
        e = s + set_size - 1
        set_amount = price_for_set_size(set_size)

        # Generate the PNG strip with no buyer details -> QR shows only numbers.
        try:
            filename = create_coupon(s, e, amount=set_amount)
            generated += 1
        except Exception as exc:
            print(f"  WARNING: Could not generate {s:04d}-{e:04d}: {exc}")
            failed += 1
            continue

        # Record the physical row in Excel (blank buyer fields).
        try:
            record_sale(
                None, None, None, s, e, sale_type="PHYSICAL",
                set_size=set_size, amount=set_amount,
            )
        except RuntimeError as exc:
            print(f"  WARNING: {exc}")
            print(f"           PNG {filename} generated but NOT tracked in Excel.")
            failed += 1
            continue

        print(f"  [{i + 1}/{num_sets}] {s:04d}-{e:04d} -> {filename}  (\u20B9{set_amount})")

    print()
    print(f"Done. {generated} set(s) generated, {failed} failed.")
    if generated:
        print(f"Numbers {start:04d}-{end:04d} are now LOCKED for physical sale.")
    return True


def prompt_delete_physical():
    """Delete a physical-sale set by S.No (only PHYSICAL rows are shown).

    Numbers are freed for re-sale (no audit log is kept), matching the
    behaviour of the normal delete flow.

    Returns True to continue the menu loop."""
    try:
        rows = list_sales(print_it=False)
    except RuntimeError as exc:
        print(f"ERROR: {exc}")
        return True

    # Keep only rows whose Type is PHYSICAL. Tolerate old 8-column rows.
    physical_rows = []
    for r in rows:
        stype = r[8] if len(r) >= 9 else "SALE"
        if stype == "PHYSICAL":
            physical_rows.append(r)

    if not physical_rows:
        print()
        print("No physical sets recorded yet. Nothing to delete.")
        return True

    print()
    print("-" * 70)
    print("PHYSICAL SETS ONLY")
    print("-" * 70)
    print(f"{'S.No':<6}{'Start':<8}{'End':<8}{'Qty':<6}{'Date Sold'}")
    print("-" * 70)
    for r in physical_rows:
        sno, name, phone, address, start, end, qty, date = r[:8]
        print(f"{sno!s:<6}{start!s:<8}{end!s:<8}{qty!s:<6}{date}")
    print("-" * 70)

    sno_raw = ask("Enter S.No to delete (or 'q' to cancel): ")
    if sno_raw is None:
        print("Delete cancelled.")
        return True

    try:
        sno = int(sno_raw)
    except ValueError:
        print("Invalid S.No. Delete cancelled.")
        return True

    # Verify the chosen S.No really is a PHYSICAL row.
    is_physical = any(int(r[0]) == sno for r in physical_rows)
    if not is_physical:
        print(f"S.No {sno} is not a physical set. Use 'Delete a sale' for normal sales.")
        return True

    confirm = ask(f"Are you sure you want to delete physical set #{sno}? (y/n): ")
    if confirm is None or confirm.lower() not in ("y", "yes"):
        print("Delete cancelled.")
        return True

    del_png = False
    png_ans = ask("Also delete the coupon PNG file? (y/n): ")
    if png_ans is not None and png_ans.lower() in ("y", "yes"):
        del_png = True

    try:
        deleted = delete_sale(sno, delete_png=del_png)
    except RuntimeError as exc:
        print(f"ERROR: {exc}")
        return True

    if deleted:
        print(f"Physical set #{sno} deleted. Those coupon numbers are now free for re-sale.")
    else:
        print(f"No physical set found with S.No #{sno}.")

    return True


def delete_all_physical(delete_pngs=False):
    """Delete every PHYSICAL row from the Excel file, keeping SALE rows.

    If delete_pngs is True, also remove the PNG files that correspond to the
    deleted physical sets. Returns the number of physical rows deleted.
    Raises RuntimeError if the Excel file is locked."""
    ensure_sales_file()
    wb = _load_sales_workbook()
    ws = wb.active

    # First pass: collect the (start, end) ranges of every PHYSICAL row,
    # plus the row indices, so we can delete them and optionally their PNGs.
    physical_ranges = []
    rows_to_delete = []

    for row_idx in range(2, ws.max_row + 1):
        stype = ws.cell(row=row_idx, column=COL_TYPE).value
        if stype == "PHYSICAL":
            start = ws.cell(row=row_idx, column=5).value
            end = ws.cell(row=row_idx, column=6).value
            physical_ranges.append((int(start), int(end)))
            rows_to_delete.append(row_idx)

    if not rows_to_delete:
        return 0

    # Delete rows from the bottom up so indices stay valid.
    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx, 1)

    # Renumber S.No for all remaining rows so they stay sequential.
    new_sno = 1
    for row_idx in range(2, ws.max_row + 1):
        cell_sno = ws.cell(row=row_idx, column=1).value
        if cell_sno is None:
            continue
        ws.cell(row=row_idx, column=1).value = new_sno
        new_sno += 1

    try:
        wb.save(SALES_FILE)
    except PermissionError:
        raise RuntimeError(
            f"Cannot save '{SALES_FILE.name}'. Please CLOSE it in "
            f"Microsoft Excel / OneDrive and try again."
        )
    finally:
        wb.close()
        _invalidate_sales_cache()

    # Optionally delete the PNG files for the deleted physical sets.
    if delete_pngs and OUTPUT_DIR.exists():
        removed = 0
        for start, end in physical_ranges:
            is_range = start != end
            num_part = f"{start:04d}-{end:04d}" if is_range else f"{start:04d}"
            # Physical coupons are saved with no buyer name -> coupon_<num>.png
            filename = f"coupon_{num_part}.png"
            png_path = OUTPUT_DIR / filename
            if png_path.exists():
                try:
                    png_path.unlink()
                    removed += 1
                except Exception:
                    pass
        if removed:
            print(f"Deleted {removed} physical-set PNG file(s) from {OUTPUT_DIR.name}.")

    return len(rows_to_delete)


def delete_physical_by_range(rng_start, rng_end, delete_pngs=False):
    """Delete every PHYSICAL row whose coupon range falls entirely within
    [rng_start, rng_end]. Normal SALE rows are left untouched.

    If delete_pngs is True, also remove the corresponding PNG files.
    Returns the number of physical rows deleted.
    Raises RuntimeError if the Excel file is locked."""
    ensure_sales_file()
    wb = _load_sales_workbook()
    ws = wb.active

    physical_ranges = []
    rows_to_delete = []

    for row_idx in range(2, ws.max_row + 1):
        stype = ws.cell(row=row_idx, column=COL_TYPE).value
        if stype != "PHYSICAL":
            continue
        start = ws.cell(row=row_idx, column=5).value
        end = ws.cell(row=row_idx, column=6).value
        if start is None or end is None:
            continue
        try:
            s, e = int(start), int(end)
        except (ValueError, TypeError):
            continue
        # The physical set must lie fully within the user's range.
        if s >= rng_start and e <= rng_end:
            physical_ranges.append((s, e))
            rows_to_delete.append(row_idx)

    if not rows_to_delete:
        return 0

    for row_idx in reversed(rows_to_delete):
        ws.delete_rows(row_idx, 1)

    # Renumber S.No for all remaining rows so they stay sequential.
    new_sno = 1
    for row_idx in range(2, ws.max_row + 1):
        cell_sno = ws.cell(row=row_idx, column=1).value
        if cell_sno is None:
            continue
        ws.cell(row=row_idx, column=1).value = new_sno
        new_sno += 1

    try:
        wb.save(SALES_FILE)
    except PermissionError:
        raise RuntimeError(
            f"Cannot save '{SALES_FILE.name}'. Please CLOSE it in "
            f"Microsoft Excel / OneDrive and try again."
        )
    finally:
        wb.close()
        _invalidate_sales_cache()

    if delete_pngs and OUTPUT_DIR.exists():
        removed = 0
        for start, end in physical_ranges:
            is_range = start != end
            num_part = f"{start:04d}-{end:04d}" if is_range else f"{start:04d}"
            filename = f"coupon_{num_part}.png"
            png_path = OUTPUT_DIR / filename
            if png_path.exists():
                try:
                    png_path.unlink()
                    removed += 1
                except Exception:
                    pass
        if removed:
            print(f"Deleted {removed} physical-set PNG file(s) from {OUTPUT_DIR.name}.")

    return len(rows_to_delete)


def prompt_delete_physical_by_range():
    """Delete physical sets whose numbers fall within a user-entered range.

    Lets you wipe a whole block at once (e.g. enter 2501-2600 to remove all
    100 sets of 1 generated in that span). Only PHYSICAL rows are affected;
    normal SALE rows inside the same span are kept.

    Returns True to continue the menu loop."""
    try:
        rows = list_sales(print_it=False)
    except RuntimeError as exc:
        print(f"ERROR: {exc}")
        return True

    physical_rows = [r for r in rows if (r[8] if len(r) >= 9 else "SALE") == "PHYSICAL"]

    if not physical_rows:
        print()
        print("No physical sets recorded. Nothing to delete.")
        return True

    print()
    print("-" * 72)
    print("DELETE PHYSICAL SETS BY RANGE")
    print("Enter a coupon-number span. Every physical set that falls fully")
    print("inside that span will be deleted. Normal sales are kept.")
    print("(type 'q' at any prompt to cancel)")
    print("-" * 72)

    rng_start = ask_int("Range Start Number: ")
    if rng_start is None:
        return True
    if rng_start < 1:
        print("ERROR: Start number must be at least 1. Cancelled.")
        return True

    rng_end = ask_int("Range End   Number: ")
    if rng_end is None:
        return True
    if rng_end < rng_start:
        print("ERROR: End number cannot be less than start number. Cancelled.")
        return True

    # Preview which physical sets will be deleted.
    matched = []
    for r in physical_rows:
        sno, name, phone, address, start, end, qty, date = r[:8]
        if int(start) >= rng_start and int(end) <= rng_end:
            matched.append(r)

    if not matched:
        print(f"No physical sets fall within {rng_start:04d}-{rng_end:04d}. Nothing to delete.")
        return True

    print()
    print(f"{len(matched)} physical set(s) match {rng_start:04d}-{rng_end:04d}:")
    print(f"{'S.No':<6}{'Start':<8}{'End':<8}{'Qty':<6}")
    print("-" * 36)
    for r in matched:
        sno, name, phone, address, start, end, qty, date = r[:8]
        print(f"{sno!s:<6}{start!s:<8}{end!s:<8}{qty!s:<6}")
    print("-" * 36)

    confirm = ask(f"Delete all {len(matched)} set(s) above? (y/n): ")
    if confirm is None or confirm.lower() not in ("y", "yes"):
        print("Delete cancelled.")
        return True

    del_png = False
    png_ans = ask("Also delete the coupon PNG files? (y/n): ")
    if png_ans is not None and png_ans.lower() in ("y", "yes"):
        del_png = True

    try:
        count = delete_physical_by_range(rng_start, rng_end, delete_pngs=del_png)
    except RuntimeError as exc:
        print(f"ERROR: {exc}")
        return True

    print(f"Deleted {count} physical set(s) in {rng_start:04d}-{rng_end:04d}.")
    print("Those coupon numbers are now free for re-sale.")
    return True


def prompt_delete_all_physical():
    """Interactive flow to delete ALL physical sets at once.
    Normal SALE rows are left untouched. Returns True to continue the menu loop."""
    try:
        rows = list_sales(print_it=False)
    except RuntimeError as exc:
        print(f"ERROR: {exc}")
        return True

    physical_count = sum(1 for r in rows if (r[8] if len(r) >= 9 else "SALE") == "PHYSICAL")

    if physical_count == 0:
        print()
        print("No physical sets recorded. Nothing to delete.")
        return True

    print()
    print(f"This will delete ALL {physical_count} physical set(s) from the Excel tracker.")
    print("Normal sales are NOT affected.")
    confirm = ask("Are you sure? This cannot be undone. (y/n): ")
    if confirm is None or confirm.lower() not in ("y", "yes"):
        print("Delete-all-physical cancelled.")
        return True

    del_png = False
    png_ans = ask("Also delete ALL physical-set PNG files? (y/n): ")
    if png_ans is not None and png_ans.lower() in ("y", "yes"):
        del_png = True

    try:
        count = delete_all_physical(delete_pngs=del_png)
    except RuntimeError as exc:
        print(f"ERROR: {exc}")
        return True

    print(f"Deleted all {count} physical set(s). Those coupon numbers are now free for re-sale.")
    return True


def view_gaps():
    """Print the coupon numbers that are NOT assigned but lie between the
    lowest and highest assigned number. Uses the union of the Excel tracker
    (assigned coupons) and the PNG files actually present on disk.

    A "gap" is any number N where:
        lowest_assigned <= N <= highest_assigned
    and N is not covered by any row in the Excel tracker.

    Returns True to continue the menu loop."""
    try:
        ranges = get_sold_ranges()
    except RuntimeError as exc:
        print(f"ERROR: {exc}")
        return True

    if not ranges:
        print()
        print("No coupons assigned yet. Nothing to compare.")
        return True

    # Build the set of every assigned coupon number from the Excel tracker.
    assigned = set()
    for s, e in ranges:
        for n in range(s, e + 1):
            assigned.add(n)

    if not assigned:
        print()
        print("No coupons assigned yet. Nothing to compare.")
        return True

    lowest = min(assigned)
    highest = max(assigned)

    # A gap number is any number in [lowest, highest] that is not assigned.
    gaps = [n for n in range(lowest, highest + 1) if n not in assigned]

    print()
    print("=" * 60)
    print("COUPON NUMBER GAPS")
    print("=" * 60)
    print(f"  Lowest assigned  : {lowest:04d}")
    print(f"  Highest assigned : {highest:04d}")
    print(f"  Total in span    : {highest - lowest + 1}")
    print(f"  Assigned         : {len(assigned)}")
    print(f"  Gaps (unassigned): {len(gaps)}")
    print("-" * 60)

    if not gaps:
        print("  No gaps. Every number from {lowest:04d} to {highest:04d} is assigned.".format(
            lowest=lowest, highest=highest
        ))
    else:
        # Collapse consecutive gap numbers into ranges for a compact view.
        # e.g. [5,6,7,9] -> "0005-0007", "0009".
        def collapse(nums):
            if not nums:
                return []
            nums = sorted(nums)
            runs = []
            run_start = nums[0]
            run_end = nums[0]
            for n in nums[1:]:
                if n == run_end + 1:
                    run_end = n
                else:
                    runs.append((run_start, run_end))
                    run_start = n
                    run_end = n
            runs.append((run_start, run_end))
            return runs

        runs = collapse(gaps)
        for s, e in runs:
            if s == e:
                print(f"  {s:04d}")
            else:
                print(f"  {s:04d}-{e:04d}  ({e - s + 1} coupons)")

    print("-" * 60)
    print()
    return True


def menu():
    """Main interactive menu loop."""
    while True:
        print()
        print("=" * 40)
        print("MENU")
        print("=" * 40)
        print("  1. New Sale")
        print("  2. Physical Set Generation")
        print("  3. Delete a sale")
        print("  4. Delete a physical set")
        print("  5. Delete ALL physical sets")
        print("  6. Delete physical sets by range")
        print("  7. View all sales")
        print("  8. Delete ALL sales")
        print("  9. Last coupon sold")
        print(" 10. View gap numbers")
        print(" 11. Exit")
        print("-" * 40)

        choice = ask("Choice: ")
        if choice is None:
            break

        choice = choice.strip()

        if choice == "1":
            if not prompt_sale():
                break
        elif choice == "2":
            prompt_physical_set()
        elif choice == "3":
            prompt_delete()
        elif choice == "4":
            prompt_delete_physical()
        elif choice == "5":
            prompt_delete_all_physical()
        elif choice == "6":
            prompt_delete_physical_by_range()
        elif choice == "7":
            try:
                list_sales()
            except RuntimeError as exc:
                print(f"ERROR: {exc}")
        elif choice == "8":
            prompt_delete_all()
        elif choice == "9":
            show_last_sale()
        elif choice == "10":
            view_gaps()
        elif choice == "11":
            break
        else:
            print("Invalid choice. Enter 1 through 11.")


# ============================================================
# MAIN
# ============================================================

def main():
    print("=" * 72)
    print("JAI BHADRA FOUNDATION - LUCKY DRAW COUPON GENERATOR")
    print("=" * 72)
    print()
    print(f"Max coupon    : {MAX_COUPON}")
    print(f"Left photo    : {LEFT_PHOTO}")
    print(f"Right photo   : {RIGHT_PHOTO}")
    print(f"Draw date     : {DRAW_DATE} ({DRAW_DAY})")
    print(f"Website       : {WEBSITE}")
    print(f"Output folder : {OUTPUT_DIR}")
    print(f"Sales tracker : {SALES_FILE}")
    print()

    if not LEFT_PHOTO.exists():
        raise FileNotFoundError(
            f"Missing required file: {LEFT_PHOTO}"
        )

    if not RIGHT_PHOTO.exists():
        raise FileNotFoundError(
            f"Missing required file: {RIGHT_PHOTO}"
        )

    try:
        ensure_sales_file()
    except RuntimeError as exc:
        print()
        print(f"ERROR: {exc}")
        print()
        return

    menu()

    print()
    print("=" * 72)
    print("DONE")
    print("=" * 72)
    print(f"Coupons saved to : {OUTPUT_DIR}")
    print(f"Sales tracked in : {SALES_FILE}")
    print()


if __name__ == "__main__":
    main()